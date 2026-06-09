"""Tests for the standardized official-grid .xlsx I/O and runner (FASE 11).

Builds a synthetic template in the official layout (anchors ``Estado inicial``
and ``#Prueba``, k blocks of six columns starting at the official base
columns) over the real N3A sample network, then exercises the reader, the
writer's resume logic and the end-to-end ``fill_grid`` engine.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.constants.grid import (
    GRID_FAMILY_OFFSET,
    GRID_K_BASE_COLUMN,
)
from src.funcs.grid import (
    GridResultsWriter,
    fill_grid,
    grid_sheet_names,
    letters_to_mask,
    read_grid_sheet,
)

SHEET = "3A-Elementos"


@pytest.fixture()
def template(tmp_path: Path) -> Path:
    """Write a minimal official-format template with two N3A test rows."""
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet.cell(row=1, column=1).value = "Estado inicial"
    sheet.cell(row=1, column=2).value = "100"
    sheet.cell(row=5, column=1).value = "#Prueba"
    for offset, (purview, mechanism) in enumerate([("ABC", "ABC"), ("ABC", "AB")]):
        sheet.cell(row=6 + offset, column=1).value = offset + 1
        sheet.cell(row=6 + offset, column=2).value = purview
        sheet.cell(row=6 + offset, column=3).value = mechanism
    path = tmp_path / "template.xlsx"
    book.save(path)
    return path


def test_letters_to_mask() -> None:
    """Letter masks map onto little-endian bit masks by alphabet position."""
    assert letters_to_mask("ACE", 5) == "10101"
    assert letters_to_mask("", 3) == "000"
    assert letters_to_mask(" AB ", 3) == "110"


def test_grid_sheet_names_filters_official_suffix(template: Path) -> None:
    """Only ``*-Elementos`` sheets count as grid sheets (case-sensitive)."""
    assert grid_sheet_names(template) == [SHEET]


def test_read_grid_sheet_parses_anchors_and_tests(template: Path) -> None:
    """The reader extracts state, page and the test rows with their masks."""
    sheet = read_grid_sheet(template, SHEET)
    assert sheet.initial_state == "100"
    assert sheet.num_nodes == 3
    assert sheet.page == "A"
    assert sheet.header_row == 5
    assert len(sheet.tests) == 2
    assert sheet.tests[0].purview_mask == "111"
    assert sheet.tests[1].mechanism_mask == "110"
    assert sheet.tests[1].row == 7


def test_read_grid_sheet_requires_anchors(tmp_path: Path) -> None:
    """A sheet without the official anchors is rejected."""
    book = Workbook()
    book.active.title = "9Z-Elementos"
    path = tmp_path / "bad.xlsx"
    book.save(path)
    with pytest.raises(ValueError):
        read_grid_sheet(path, "9Z-Elementos")


def test_writer_missing_ks_and_resume(template: Path, tmp_path: Path) -> None:
    """missing_ks reflects already-written loss cells (resume contract)."""
    output = tmp_path / "out.xlsx"
    writer = GridResultsWriter(template, output)
    assert writer.missing_ks(SHEET, 6, "QNodes") == (2, 3, 4, 5)

    writer.write_result(SHEET, 6, "QNodes", 3, "P", 0.5, 1.0)
    writer.save()

    resumed = GridResultsWriter(template, output)
    assert resumed.missing_ks(SHEET, 6, "QNodes") == (2, 4, 5)
    assert resumed.missing_ks(SHEET, 6, "Geometric") == (2, 3, 4, 5)


def test_fill_grid_end_to_end_n3(template: Path, tmp_path: Path) -> None:
    """fill_grid completes every (row, family, k) cell and spares the template."""
    output = tmp_path / "results.xlsx"
    fill_grid(template, output, k_values=(2, 3), progress=lambda _: None)

    book = load_workbook(output)
    sheet = book[SHEET]
    for row in (6, 7):
        for family, offset in GRID_FAMILY_OFFSET.items():
            for k in (2, 3):
                base = GRID_K_BASE_COLUMN[k] + offset
                assert sheet.cell(row=row, column=base).value, (row, family, k)
                assert sheet.cell(row=row, column=base + 1).value is not None
                assert sheet.cell(row=row, column=base + 2).value is not None

    pristine = load_workbook(template)[SHEET]
    assert pristine.cell(row=6, column=GRID_K_BASE_COLUMN[2]).value is None
