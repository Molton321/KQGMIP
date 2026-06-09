"""Fill the official evaluation grid (DatosPruebas2026_1.xlsx) with KQNodes/KGeoMIP.

The official workbook ``data/results/DatosPruebas2026_1.xlsx`` is a *template*: one
sheet per network (10A, 15B, 20A, 22A, 25A), each listing ~50 subsystem test rows
by their purview (``Alcance``) and mechanism (``Mecanismo``) letter masks, with
empty ``Partición / Pérdida / Tiempo`` cells for the ``QNodes`` and ``Geometric``
strategies across k = 2, 3, 4, 5.

This script reads that template, runs ``KQNodes`` and ``KGeoMIP`` for every
(row, k) and writes the results to a **separate** workbook
(``data/results/Resultados_DatosPruebas2026_1.xlsx``) so the official template is
never modified. The convention was cross-validated against ``Pruebas_Metodo2.xlsx``
(known GeoMIP/PyPhi losses): the candidate state is the sheet's ``Estado inicial``,
the background condition is the full system (all-ones), and the subsystem is carved
by the purview/mechanism letter masks.

Progress is saved after each row so a long run can be interrupted and resumed (the
output is reloaded and rows already filled are skipped).

    uv run scripts/fill_official_grid.py --sheets 10A-Elementos
    uv run scripts/fill_official_grid.py            # every sheet (slow on n>=20)
"""

import argparse
import contextlib
import io
import string
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# isort: split
from openpyxl import load_workbook

from src.constants.base import PATH_RESULTS
from src.controllers.manager import Manager
from src.funcs.runner import build_strategy
from src.models.base.application import application

TEMPLATE = PATH_RESULTS / "DatosPruebas2026_1.xlsx"
OUTPUT = PATH_RESULTS / "Resultados_DatosPruebas2026_1.xlsx"

K_VALUES = (2, 3, 4, 5)
"""The k-partition orders the grid evaluates per subsystem."""

K_BASE_COLUMN = {2: 4, 3: 10, 4: 16, 5: 22}
"""1-based openpyxl column of the QNodes 'Partición' cell for each k block.

Each block is six columns: QNodes (Partición, Pérdida, Tiempo) then Geometric
(Partición, Pérdida, Tiempo). Geometric starts three columns after QNodes.
"""

STRATEGY_KEY = {"QNodes": "KQNodes", "Geometric": "KGeoMIP"}
"""Maps the grid's column family to the project's k-partition strategy."""


def letters_to_mask(letters: str, n: int) -> str:
    """Convert a purview/mechanism letter string (``"ACEGI"``) to an n-bit mask."""
    bits = ["0"] * n
    for char in str(letters).strip():
        bits[string.ascii_uppercase.index(char)] = "1"
    return "".join(bits)


def _find_anchors(sheet) -> tuple[str, int]:
    """Return ``(initial_state, header_row)`` for a grid sheet (1-based row)."""
    initial_state = ""
    header_row = 0
    for row in range(1, 12):
        label = sheet.cell(row=row, column=1).value
        if label == "Estado inicial":
            initial_state = str(sheet.cell(row=row, column=2).value).strip()
        elif label == "#Prueba":
            header_row = row
            break
    if not initial_state or not header_row:
        raise ValueError("sheet is missing the 'Estado inicial' or '#Prueba' anchor")
    return initial_state, header_row


def _run_cell(tpm, state: str, key: str, k: int, condition: str, purview: str, mechanism: str):
    """Run one strategy and return ``(partition_str, loss, seconds)``."""
    start = time.perf_counter()
    with contextlib.redirect_stdout(io.StringIO()):
        analyzer = build_strategy(key, tpm, state, k, "spectral")
        solution = analyzer.apply_strategy(condition, purview, mechanism)
    elapsed = time.perf_counter() - start
    return solution.partition, round(float(solution.loss), 8), round(elapsed, 4)


def fill_sheet(template_book, output_book, sheet_name: str) -> None:
    """Fill every (row, k, strategy) cell of one sheet, saving after each row."""
    src = template_book[sheet_name]
    dst = output_book[sheet_name]
    state, header_row = _find_anchors(src)
    n = len(state)
    condition = "1" * n
    page = "".join(char for char in sheet_name if char.isalpha())[:1]
    application.set_sample_network_page(page)
    tpm = Manager(state).load_network()

    row = header_row + 1
    filled_rows = 0
    while True:
        purview_letters = src.cell(row=row, column=2).value
        mechanism_letters = src.cell(row=row, column=3).value
        if purview_letters in (None, "") or mechanism_letters in (None, ""):
            break
        purview = letters_to_mask(purview_letters, n)
        mechanism = letters_to_mask(mechanism_letters, n)
        for k in K_VALUES:
            base = K_BASE_COLUMN[k]
            for offset, family in ((0, "QNodes"), (3, "Geometric")):
                loss_col = base + offset + 1
                if dst.cell(row=row, column=loss_col).value not in (None, ""):
                    continue
                partition, loss, seconds = _run_cell(
                    tpm, state, STRATEGY_KEY[family], k, condition, purview, mechanism
                )
                dst.cell(row=row, column=base + offset).value = partition
                dst.cell(row=row, column=loss_col).value = loss
                dst.cell(row=row, column=base + offset + 2).value = seconds
        output_book.save(OUTPUT)
        filled_rows += 1
        print(f"  {sheet_name} row {row - header_row}: {purview_letters}|{mechanism_letters} done")
        row += 1
    print(f"{sheet_name}: {filled_rows} rows filled (n={n}, state={state}).")


def main() -> None:
    """Parse options, fill the requested sheets and save the results workbook."""
    parser = argparse.ArgumentParser(description="Fill the official k-partition grid.")
    parser.add_argument(
        "--sheets",
        nargs="*",
        default=None,
        help="Sheet names to fill (default: all *-Elementos sheets).",
    )
    args = parser.parse_args()
    application.disable_profiling()

    template_book = load_workbook(TEMPLATE)
    output_book = load_workbook(OUTPUT) if OUTPUT.exists() else load_workbook(TEMPLATE)

    sheets = args.sheets or [s for s in template_book.sheetnames if s.strip().endswith("Elementos")]
    for sheet_name in sheets:
        print(f"=== {sheet_name} ===")
        fill_sheet(template_book, output_book, sheet_name)
    output_book.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")


if __name__ == "__main__":
    main()
