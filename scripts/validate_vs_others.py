"""Cross-validation against external sources (FASE 11, Invariante 3/4).

Validates the project's results against every independent source available:

1. **Original project from the professor** (``.core/core_00/GeoMIP``): the TPM
   samples are verified byte-identical and the rows of her
   ``resultados_Geometric.xlsx`` (N15A, continuous TPM) are reproduced with our
   ``GeometricSIA`` within float32 tolerance.
2. **Third-party KQNodes/QNodes CSVs** (``data/results_others/{kqnodes,qnodes}``):
   per-(network, k, subsystem) loss comparison against our filled official
   table (``Resultados_DatosPruebas2026_1.xlsx``). k=2 must tie exactly
   (both implementations search the same bipartition space); for k>=3 lower
   delta is better.
3. **Third-party workbook** (``data/results_others/DatosPruebas2026_1.xlsx``):
   their k=2 losses are compared against the validated exact optimum; values
   *below* the exhaustive minimum prove a different loss convention (nothing
   can beat the exact minimum of the same function), so that workbook is
   reported as non-comparable rather than better/worse.

    uv run scripts/validate_vs_others.py                 # fast checks (1-3)
    uv run scripts/validate_vs_others.py --docente-rows 6
"""

import argparse
import contextlib
import io
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# isort: split
import pandas as pd
from openpyxl import load_workbook

from src.constants.base import PATH_RESULTS, PROJECT_ROOT
from src.constants.grid import GRID_FAMILY_OFFSET, GRID_K_BASE_COLUMN, GRID_K_VALUES
from src.controllers.manager import Manager
from src.controllers.strategies.geometric import GeometricSIA
from src.models.base.application import application

PATH_OTHERS = PROJECT_ROOT / "data" / "results_others"
PATH_CORE = PROJECT_ROOT / ".core" / "core_00" / "GeoMIP"
OURS_XLSX = PATH_RESULTS / "Resultados_DatosPruebas2026_1.xlsx"

LOSS_EPS = 1e-6
"""Comparison tolerance: above the 8-decimal rounding of our stored losses."""


def load_ours() -> pd.DataFrame:
    """Read our filled official table into tidy (net, k, family, subsystem, loss) rows."""
    book = load_workbook(OURS_XLSX, read_only=True)
    rows = []
    for name in book.sheetnames:
        clean = name.strip()
        if not clean.endswith("Elementos"):
            continue
        net = "N" + clean.replace("-Elementos", "")
        sheet = book[name]
        header = next(r for r in range(1, 12) if sheet.cell(row=r, column=1).value == "#Prueba")
        row = header + 1
        while True:
            purview = sheet.cell(row=row, column=2).value
            mechanism = sheet.cell(row=row, column=3).value
            if purview in (None, "") or mechanism in (None, ""):
                break
            for k in GRID_K_VALUES:
                for family, offset in GRID_FAMILY_OFFSET.items():
                    loss = sheet.cell(row=row, column=GRID_K_BASE_COLUMN[k] + offset + 1).value
                    if loss not in (None, ""):
                        rows.append(
                            dict(
                                net=net,
                                k=k,
                                purview=str(purview).strip(),
                                mechanism=str(mechanism).strip(),
                                family=family,
                                loss_ours=float(loss),
                            )
                        )
            row += 1
    book.close()
    return pd.DataFrame(rows)


def load_theirs_kqnodes() -> pd.DataFrame:
    """Read the third-party KQNodes CSVs (one file per network and k)."""
    rows = []
    for path in sorted((PATH_OTHERS / "kqnodes").glob("resultado__*.csv")):
        match = re.match(r"resultado__N(\d+)_([A-Z])_(\d)\.csv", path.name)
        if not match:
            continue
        net = f"N{match.group(1)}{match.group(2)}"
        k = int(match.group(3))
        for _, record in pd.read_csv(path).iterrows():
            rows.append(
                dict(
                    net=net,
                    k=k,
                    purview=record["Alcance"].strip(),
                    mechanism=record["Mecanismo"].strip(),
                    loss_other=float(record["Pérdida (φ)"]),
                )
            )
    return pd.DataFrame(rows)


def load_theirs_qnodes() -> pd.DataFrame:
    """Read the third-party legacy QNodes (k=2) CSVs."""
    rows = []
    for path in sorted((PATH_OTHERS / "qnodes").glob("resultado__*.csv")):
        match = re.match(r"resultado__N(\d+)_([A-Z])\.csv", path.name)
        if not match:
            continue
        net = f"N{match.group(1)}{match.group(2)}"
        for _, record in pd.read_csv(path).iterrows():
            rows.append(
                dict(
                    net=net,
                    k=2,
                    purview=record["Alcance"].strip(),
                    mechanism=record["Mecanismo"].strip(),
                    loss_other=float(record["Pérdida (φ)"]),
                )
            )
    return pd.DataFrame(rows)


def compare_with_csv_team() -> None:
    """Win/tie/loss table of our strategies vs the CSV third party, per (net, k)."""
    ours = load_ours()
    theirs = pd.concat([load_theirs_kqnodes(), load_theirs_qnodes()], ignore_index=True)
    theirs = theirs.drop_duplicates(subset=["net", "k", "purview", "mechanism"], keep="first")
    merged = ours.merge(theirs, on=["net", "k", "purview", "mechanism"])
    merged["diff"] = merged["loss_ours"] - merged["loss_other"]

    print("== Nuestras estrategias vs terceros (CSV kqnodes/qnodes); δ menor = mejor ==")
    for family, group in merged.groupby("family"):
        wins = int((group["diff"] < -LOSS_EPS).sum())
        ties = int((group["diff"].abs() <= LOSS_EPS).sum())
        losses = int((group["diff"] > LOSS_EPS).sum())
        print(f"  {family}: {len(group)} casos -> ganamos {wins}, empate {ties}, perdemos {losses}")
        k2 = group[group.k == 2]
        k2_ties = int((k2["diff"].abs() <= LOSS_EPS).sum())
        print(f"    k=2: {k2_ties}/{len(k2)} empates exactos (misma convención y mismo óptimo)")
        if losses:
            print(group[group["diff"] > LOSS_EPS].to_string(index=False))


def reproduce_docente(num_rows: int) -> None:
    """Reproduce the professor's ``resultados_Geometric.xlsx`` rows with GeometricSIA."""
    sample_ours = PROJECT_ROOT / "data" / "samples" / "N15A.csv"
    sample_core = PATH_CORE / "data" / "samples" / "N15A.csv"
    identical = sample_ours.read_bytes() == sample_core.read_bytes()
    print(f"== Proyecto original (.core/core_00): TPM N15A idéntica byte a byte: {identical} ==")

    book = load_workbook(PATH_CORE / "results" / "resultados_Geometric.xlsx", read_only=True)
    sheet = book["Sheet1"]
    cases = []
    for row in range(2, 2 + num_rows):
        purview = str(sheet.cell(row=row, column=2).value)
        mechanism = str(sheet.cell(row=row, column=3).value)
        loss = float(str(sheet.cell(row=row, column=5).value).replace(",", "."))
        cases.append((purview, mechanism, loss))
    book.close()

    application.set_sample_network_page("A")
    state = "1" + "0" * 14
    tpm = Manager(state).load_network()
    condition = "1" * 15
    matched = 0
    for purview, mechanism, expected in cases:
        with contextlib.redirect_stdout(io.StringIO()):
            solution = GeometricSIA(tpm, state).apply_strategy(condition, purview, mechanism)
        ok = abs(float(solution.loss) - expected) < 1e-4
        matched += ok
        print(
            f"  {purview}|{mechanism}: docente={expected:.10f} "
            f"nuestra={float(solution.loss):.10f} -> {'OK' if ok else 'DIFIERE'}"
        )
    print(f"  Reproducidas {matched}/{len(cases)} filas (tolerancia float32 sobre TPM continua).")


def check_others_xlsx() -> None:
    """Show the third-party workbook uses a different (non-comparable) loss."""
    path = PATH_OTHERS / "DatosPruebas2026_1.xlsx"
    if not path.exists():
        print("== Workbook de terceros no presente; se omite ==")
        return
    book = load_workbook(path, read_only=True)
    sheet = book["10A-Elementos"]
    rows = []
    row = 6
    while True:
        purview = sheet.cell(row=row, column=2).value
        mechanism = sheet.cell(row=row, column=3).value
        if purview in (None, "") or mechanism in (None, ""):
            break
        rows.append(
            dict(
                purview=str(purview).strip(),
                mechanism=str(mechanism).strip(),
                qn=pd.to_numeric(sheet.cell(row=row, column=5).value, errors="coerce"),
                geo=pd.to_numeric(sheet.cell(row=row, column=8).value, errors="coerce"),
            )
        )
        row += 1
    book.close()

    ours = load_ours()
    optimum = ours[(ours.net == "N10A") & (ours.k == 2) & (ours.family == "QNodes")]
    merged = pd.DataFrame(rows).merge(
        optimum[["purview", "mechanism", "loss_ours"]], on=["purview", "mechanism"]
    )
    print("== Workbook xlsx de terceros vs óptimo exacto oficial (N10A, k=2) ==")
    for column, label in (("qn", "QNodes"), ("geo", "Geometric")):
        valid = merged.dropna(subset=[column])
        below = int((valid[column] < valid.loss_ours - LOSS_EPS).sum())
        equal = int(((valid[column] - valid.loss_ours).abs() <= LOSS_EPS).sum())
        print(
            f"  {label}: {len(valid)} filas -> {below} por DEBAJO del mínimo exhaustivo, "
            f"{equal} iguales. Valores bajo el mínimo exacto son imposibles con la δ oficial: "
            "métrica distinta, workbook NO comparable."
        )


def main() -> None:
    """Run the requested cross-validations (fast set by default)."""
    parser = argparse.ArgumentParser(description="Cross-validation vs external sources.")
    parser.add_argument("--docente-rows", type=int, default=6, help="Rows of resultados_Geometric to reproduce.")
    args = parser.parse_args()
    application.disable_profiling()

    compare_with_csv_team()
    print()
    reproduce_docente(args.docente_rows)
    print()
    check_others_xlsx()


if __name__ == "__main__":
    main()
