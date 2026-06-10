"""Standardized reader/writer/runner for grid workbooks,
which have a fixed format and are used for batch evaluation of the core strategies.
"""

import contextlib
import io
import string
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.constants.grid import (
    GRID_FAMILY_OFFSET,
    GRID_FAMILY_STRATEGY,
    GRID_HEADER_LABEL,
    GRID_K_BASE_COLUMN,
    GRID_K_VALUES,
    GRID_MECHANISM_COLUMN,
    GRID_PURVIEW_COLUMN,
    GRID_SHEET_SUFFIX,
    GRID_STATE_LABEL,
)
from src.controllers.manager import Manager
from src.funcs.runner import build_strategy
from src.models.base.application import application


@dataclass(frozen=True)
class GridTest:
    """One test row of a grid sheet: purview/mechanism letter masks + bit masks."""

    row: int
    purview_letters: str
    mechanism_letters: str
    purview_mask: str
    mechanism_mask: str


@dataclass(frozen=True)
class GridSheet:
    """One parsed ``*-Elementos`` sheet: network metadata plus its test rows."""

    name: str
    initial_state: str
    page: str
    header_row: int
    tests: tuple[GridTest, ...]

    @property
    def num_nodes(self) -> int:
        """Network size n (length of the candidate initial state)."""
        return len(self.initial_state)


def letters_to_mask(letters: str, num_nodes: int) -> str:
    """Convert a purview/mechanism letter string (``"ACEGI"``) to an n-bit mask."""
    bits = ["0"] * num_nodes
    for char in str(letters).strip():
        bits[string.ascii_uppercase.index(char)] = "1"
    return "".join(bits)


def grid_sheet_names(path: Path) -> list[str]:
    """Return the ``*-Elementos`` sheet names of a grid workbook."""
    book = load_workbook(path, read_only=True)
    try:
        return [
            name for name in book.sheetnames if name.strip().endswith(GRID_SHEET_SUFFIX)
        ]
    finally:
        book.close()


def read_grid_sheet(path: Path, sheet_name: str) -> GridSheet:
    """Parse one grid sheet into its standardized representation."""
    book = load_workbook(path, read_only=True)
    try:
        sheet = book[sheet_name]
        initial_state = ""
        header_row = 0
        for row in range(1, 12):
            label = sheet.cell(row=row, column=1).value
            if label == GRID_STATE_LABEL:
                initial_state = str(
                    sheet.cell(row=row, column=GRID_PURVIEW_COLUMN).value
                ).strip()
            elif label == GRID_HEADER_LABEL:
                header_row = row
                break
        if not initial_state or not header_row:
            raise ValueError(
                f"La hoja '{sheet_name}' no tiene las anclas "
                f"'{GRID_STATE_LABEL}' / '{GRID_HEADER_LABEL}'."
            )

        num_nodes = len(initial_state)
        tests: list[GridTest] = []
        row = header_row + 1
        while True:
            purview_letters = sheet.cell(row=row, column=GRID_PURVIEW_COLUMN).value
            mechanism_letters = sheet.cell(row=row, column=GRID_MECHANISM_COLUMN).value
            if purview_letters in (None, "") or mechanism_letters in (None, ""):
                break
            tests.append(
                GridTest(
                    row=row,
                    purview_letters=str(purview_letters).strip(),
                    mechanism_letters=str(mechanism_letters).strip(),
                    purview_mask=letters_to_mask(str(purview_letters), num_nodes),
                    mechanism_mask=letters_to_mask(str(mechanism_letters), num_nodes),
                )
            )
            row += 1

        page = next(char for char in sheet_name if char.isalpha())
        return GridSheet(
            name=sheet_name,
            initial_state=initial_state,
            page=page,
            header_row=header_row,
            tests=tuple(tests),
        )
    finally:
        book.close()


class GridResultsWriter:
    """Incremental writer for the results workbook,
    with utilities to find missing k values and write results in the right cells.
    """

    def __init__(self, template_path: Path, output_path: Path) -> None:
        """Load the output workbook (or seed it from the template)."""
        self.output_path = output_path
        self.book = load_workbook(
            output_path if output_path.exists() else template_path
        )

    def missing_ks(self, sheet_name: str, row: int, family: str) -> tuple[int, ...]:
        """Return the k values whose loss cell is still empty for this row/family."""
        sheet = self.book[sheet_name]
        offset = GRID_FAMILY_OFFSET[family]
        return tuple(
            k
            for k in GRID_K_VALUES
            if sheet.cell(row=row, column=GRID_K_BASE_COLUMN[k] + offset + 1).value
            in (None, "")
        )

    def write_result(
        self,
        sheet_name: str,
        row: int,
        family: str,
        k: int,
        partition: str,
        loss: float,
        seconds: float,
    ) -> None:
        """Write one ``Partición / Pérdida / Tiempo`` cell group."""
        sheet = self.book[sheet_name]
        base = GRID_K_BASE_COLUMN[k] + GRID_FAMILY_OFFSET[family]
        sheet.cell(row=row, column=base).value = partition
        sheet.cell(row=row, column=base + 1).value = round(float(loss), 8)
        sheet.cell(row=row, column=base + 2).value = round(float(seconds), 4)

    def save(self) -> None:
        """Persist the workbook to the output path."""
        self.book.save(self.output_path)


def read_grid_results(path: Path) -> list[dict]:
    """Read a filled grid workbook into tidy result rows.

    Each row is one (test, strategy, k) cell:
    ``{red, n, prueba, alcance, mecanismo, estrategia, k, particion, perdida, tiempo}``.
    Empty cells are skipped, so partially filled workbooks are readable too.
    """
    rows: list[dict] = []
    for sheet_name in grid_sheet_names(path):
        sheet = read_grid_sheet(path, sheet_name)
        red = "N" + sheet_name.strip().replace(f"-{GRID_SHEET_SUFFIX}", "")
        book = load_workbook(path, read_only=True)
        try:
            cells = book[sheet_name]
            for test in sheet.tests:
                for k in GRID_K_VALUES:
                    for family, offset in GRID_FAMILY_OFFSET.items():
                        base = GRID_K_BASE_COLUMN[k] + offset
                        loss = cells.cell(row=test.row, column=base + 1).value
                        if loss in (None, ""):
                            continue
                        rows.append(
                            {
                                "red": red,
                                "n": sheet.num_nodes,
                                "prueba": test.row - sheet.header_row,
                                "alcance": test.purview_letters,
                                "mecanismo": test.mechanism_letters,
                                "estrategia": family,
                                "k": k,
                                "particion": str(
                                    cells.cell(row=test.row, column=base).value or ""
                                ),
                                "perdida": float(loss),
                                "tiempo": float(
                                    cells.cell(row=test.row, column=base + 2).value or 0
                                ),
                            }
                        )
        finally:
            book.close()
    return rows


def format_results_text(rows: list[dict], max_rows: int | None = None) -> str:
    """Render tidy result rows as an aligned, console-friendly table.

    One block per network: a per-(strategy, k) summary first (count, mean and
    max loss, mean time) and then the detailed rows with partitions compacted
    to one line. Designed to be readable without knowing the project.
    """
    if not rows:
        return "Sin resultados: el archivo no tiene celdas llenas con el formato estándar."

    lines: list[str] = []
    networks = sorted({r["red"] for r in rows}, key=lambda red: (len(red), red))
    for red in networks:
        net_rows = [r for r in rows if r["red"] == red]
        lines.append("")
        lines.append(f"━━━ {red} (n={net_rows[0]['n']}, {len(net_rows)} resultados) ━━━")
        lines.append("")
        lines.append("  Resumen por estrategia y k:")
        lines.append("  Estrategia  k   casos   δ media      δ máx        t medio (s)")
        for family in sorted({r["estrategia"] for r in net_rows}):
            for k in sorted({r["k"] for r in net_rows}):
                cell = [r for r in net_rows if r["estrategia"] == family and r["k"] == k]
                if not cell:
                    continue
                losses = [r["perdida"] for r in cell]
                times = [r["tiempo"] for r in cell]
                lines.append(
                    f"  {family:<10}  {k}   {len(cell):>5}   "
                    f"{sum(losses) / len(losses):<10.6f}  {max(losses):<10.6f}  "
                    f"{sum(times) / len(times):>10.3f}"
                )
        lines.append("")
        lines.append(
            "  #   k  Estrategia  Alcance/Mecanismo            δ (pérdida)   t (s)      Partición"
        )
        detail = sorted(net_rows, key=lambda r: (r["prueba"], r["k"], r["estrategia"]))
        if max_rows is not None:
            detail = detail[:max_rows]
        for r in detail:
            subsystem = f"{r['alcance']}|{r['mecanismo']}"
            if len(subsystem) > 27:
                subsystem = subsystem[:24] + "..."
            partition = " · ".join(p.strip() for p in r["particion"].splitlines())
            if len(partition) > 48:
                partition = partition[:45] + "..."
            lines.append(
                f"  {r['prueba']:<3} {r['k']}  {r['estrategia']:<10}  {subsystem:<27}  "
                f"{r['perdida']:<12.8f}  {r['tiempo']:<9.4f}  {partition}"
            )
        if max_rows is not None and len(net_rows) > max_rows:
            lines.append(f"  … {len(net_rows) - max_rows} filas más (use --completo para verlas).")
    return "\n".join(lines)


def fill_grid(
    template_path: Path,
    output_path: Path,
    sheet_names: list[str] | None = None,
    k_values: tuple[int, ...] = GRID_K_VALUES,
    progress: Callable[[str], None] = print,
) -> None:
    """Run KQNodes + KGeoMIP over a grid workbook and fill the results copy."""
    writer = GridResultsWriter(template_path, output_path)
    names = sheet_names or grid_sheet_names(template_path)

    for sheet_name in names:
        sheet = read_grid_sheet(template_path, sheet_name)
        application.set_sample_network_page(sheet.page)
        condition = "1" * sheet.num_nodes
        tpm = Manager(sheet.initial_state).load_network()
        progress(
            f"=== {sheet_name} (n={sheet.num_nodes}, estado={sheet.initial_state}) ==="
        )

        for test in sheet.tests:
            for family, strategy_name in GRID_FAMILY_STRATEGY.items():
                missing = tuple(
                    k
                    for k in writer.missing_ks(sheet_name, test.row, family)
                    if k in k_values
                )
                if not missing:
                    continue
                with contextlib.redirect_stdout(io.StringIO()):
                    analyzer = build_strategy(
                        strategy_name, tpm, sheet.initial_state, missing[0], "spectral"
                    )
                    solutions = analyzer.apply_strategy_for_ks(
                        condition, test.purview_mask, test.mechanism_mask, missing
                    )
                for k, solution in solutions.items():
                    writer.write_result(
                        sheet_name,
                        test.row,
                        family,
                        k,
                        solution.partition,
                        float(solution.loss),
                        float(solution.execution_time),
                    )
            writer.save()
            progress(
                f"  {sheet_name} fila {test.row - sheet.header_row}: "
                f"{test.purview_letters}|{test.mechanism_letters} lista"
            )
    writer.save()
    progress(f"Resultados guardados en {output_path}")
